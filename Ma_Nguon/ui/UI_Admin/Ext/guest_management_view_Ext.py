
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QLineEdit, QComboBox, QMessageBox, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QSizePolicy, QDateEdit, QScrollArea
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QColor

from models.guest import Guest
from ui.UI_Common.custom_popup import show_success, show_error, show_warning, show_info, ask_question, ask_danger
from ui.UI_Admin.generated.ui_dialog_them_khach_thue_UI import Ui_DialogFormKhachThue



_stat_card_counter = 0

class StatCard(QFrame):
    def __init__(self, icon: str, value: str, label: str, color: str, parent=None):
        super().__init__(parent)
        global _stat_card_counter
        _stat_card_counter += 1
        obj_name = f"guestStatCard_{_stat_card_counter}"
        self.setObjectName(obj_name)
        self.setStyleSheet(
            f"QFrame#{obj_name} {{ background-color: white; border: 1px solid #e2e8f0;"
            f" border-radius: 10px; padding: 12px; }}"
            f" QFrame#{obj_name} QLabel {{ border: none; }}")
        layout = QHBoxLayout(self)
        layout.setSpacing(10)
        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet(f"font-size: 24px; background: {color}; border-radius: 8px; padding: 8px; border: none;")
        icon_lbl.setFixedSize(44, 44)
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_lbl)
        info = QVBoxLayout()
        info.setSpacing(2)
        val_lbl = QLabel(value)
        val_lbl.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        val_lbl.setStyleSheet("color: #2d3748; border: none;")
        info.addWidget(val_lbl)
        desc_lbl = QLabel(label)
        desc_lbl.setStyleSheet("color: #718096; font-size: 11px; border: none;")
        info.addWidget(desc_lbl)
        layout.addLayout(info)
        self.val_lbl = val_lbl


# ── Guest Form Dialog ─────────────────────────────────────
class GuestFormDialog(QDialog):
    """Dialog thêm/sửa khách thuê — calendar picker + full validation."""

    def __init__(self, guest: Guest = None, rooms=None, parent=None):
        super().__init__(parent)
        self.guest = guest
        self.ui = Ui_DialogFormKhachThue()
        self.ui.setupUi(self)

        # ── Replace date QLineEdits with QDateEdit (calendar popup) ──
        today = QDate.currentDate()
        self._replace_with_date_edit('inpDob', QDate(2000, 1, 1))
        self._replace_with_date_edit('inpStartDate', today)
        self._replace_with_date_edit('inpEndDate', today.addYears(1))

        # Populate rooms combo
        if rooms:
            self.ui.cbRoom.clear()
            self.ui.cbRoom.addItem("— Chọn phòng —")
            for r in rooms:
                status_vn = {'occupied': 'Đã thuê', 'available': 'Còn trống'}.get(r.status, r.status)
                self.ui.cbRoom.addItem(f"{r.room_number} ({status_vn})", r.id)

        if guest:
            self.setWindowTitle("✏️ Cập nhật khách thuê")
            self._fill_form(guest)

        self.ui.btnCancel.clicked.connect(self.reject)
        self.ui.btnSave.clicked.connect(self._on_save)

    def _replace_with_date_edit(self, attr_name: str, default_date: QDate = None):
        """Thay QLineEdit bằng QDateEdit có popup lịch."""
        old = getattr(self.ui, attr_name, None)
        if not old:
            return

        de = QDateEdit(old.parentWidget())
        de.setCalendarPopup(True)
        de.setDisplayFormat("dd/MM/yyyy")
        de.setDate(default_date or QDate.currentDate())
        de.setStyleSheet(
            "QDateEdit { background-color: white; border: 1px solid #e2e8f0; "
            "border-radius: 6px; padding: 8px; font-size: 13px; color: #1a202c; }"
            "QDateEdit:focus { border: 1px solid #3182ce; }"
            "QDateEdit::drop-down { border: none; padding-right: 8px; }")
        de.setMinimumHeight(old.minimumHeight() if old.minimumHeight() > 0 else 36)

        # Find the sub-layout (vboxlayoutN) that contains this widget
        # Each field group is: vboxlayoutN -> [label, input]
        # They're added to formGrid, then formGrid -> vboxlayout -> scrollContents
        form_grid = self.ui.formGrid
        for i in range(form_grid.count()):
            item = form_grid.itemAt(i)
            if item and item.layout():
                sub_layout = item.layout()
                idx = sub_layout.indexOf(old)
                if idx >= 0:
                    sub_layout.replaceWidget(old, de)
                    old.hide()
                    old.deleteLater()
                    setattr(self.ui, attr_name, de)
                    return

    def _fill_form(self, guest: Guest):
        self.ui.inpName.setText(guest.full_name)
        self.ui.inpPhone.setText(str(guest.phone))
        self.ui.inpIdCard.setText(str(guest.id_card))
        self.ui.inpAddress.setText(guest.address)

    def _on_save(self):
        name = self.ui.inpName.text().strip()
        phone = self.ui.inpPhone.text().strip()
        id_card = self.ui.inpIdCard.text().strip()
        address = self.ui.inpAddress.text().strip()
        price_str = self.ui.inpPrice.text().strip().replace('.', '').replace(',', '')

        # ── Bắt buộc nhập đủ tất cả ──
        if not name:
            show_warning(self, "Lỗi", "Vui lòng nhập họ tên"); return
        if not phone:
            show_warning(self, "Lỗi", "Vui lòng nhập số điện thoại"); return
        if not phone.isdigit() or len(phone) != 10:
            show_warning(self, "Lỗi", "Số điện thoại phải đủ 10 chữ số"); return
        if not id_card:
            show_warning(self, "Lỗi", "Vui lòng nhập CCCD/CMND"); return
        if not id_card.isdigit() or len(id_card) != 12:
            show_warning(self, "Lỗi", "CCCD/CMND phải đủ 12 chữ số"); return
        if not address:
            show_warning(self, "Lỗi", "Vui lòng nhập địa chỉ"); return
        if self.ui.cbRoom.currentIndex() == 0:
            show_warning(self, "Lỗi", "Vui lòng chọn phòng"); return
        if not price_str:
            show_warning(self, "Lỗi", "Vui lòng nhập giá thuê"); return

        dob = self.ui.inpDob.date().toString("yyyy-MM-dd")
        start_date = self.ui.inpStartDate.date().toString("yyyy-MM-dd")
        end_date = self.ui.inpEndDate.date().toString("yyyy-MM-dd")

        self.result_data = {
            'full_name': name, 'phone': phone, 'id_card': id_card,
            'address': address, 'dob': dob,
            'room_id': self.ui.cbRoom.currentData(),
            'price': int(price_str),
            'start_date': start_date, 'end_date': end_date,
        }
        self.accept()


# ── Guest Detail Dialog ───────────────────────────────────
class GuestDetailDialog(QDialog):
    """Dialog xem chi tiết khách thuê — dạng hồ sơ."""
    terminate_requested = pyqtSignal(int)

    def __init__(self, guest: Guest, contract_info: dict = None, parent=None):
        super().__init__(parent)
        self.guest = guest
        self.contract_info = contract_info or {}
        self.setWindowTitle(f"Hồ sơ — {guest.full_name}")
        self.setFixedSize(480, 600)
        self.setStyleSheet("QDialog { background-color: white; } QLabel { color: #2d3748; }")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        self._build_ui()

    def _box(self, label: str, value: str, color: str = None) -> QFrame:
        box = QFrame()
        box.setStyleSheet("QFrame{background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px;}")
        lay = QVBoxLayout(box)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(3)
        lbl = QLabel(label)
        lbl.setStyleSheet("color:#a0aec0;font-size:10px;font-weight:bold;border:none;background:transparent;")
        lay.addWidget(lbl)
        if color:
            val = QLabel(f"● {value}")
            val.setStyleSheet(f"color:{color};font-size:13px;font-weight:bold;border:none;background:transparent;")
        else:
            val = QLabel(str(value) if value else "—")
            val.setStyleSheet("color:#2d3748;font-size:13px;font-weight:bold;border:none;background:transparent;")
        val.setWordWrap(True)
        lay.addWidget(val)
        return box

    def _section_label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet("color:#718096;font-size:11px;font-weight:bold;letter-spacing:1px;")
        return lbl

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # Scroll area wrapping everything
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:white;}")

        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(10)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header
        h = QHBoxLayout()
        t = QLabel("👤 Hồ sơ khách thuê")
        t.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        h.addWidget(t)
        h.addStretch()
        bx = QPushButton("✕")
        bx.setFixedSize(28, 28)
        bx.setStyleSheet("QPushButton{background:transparent;border:none;font-size:16px;color:#a0aec0;}"
                          "QPushButton:hover{color:#e53e3e;}")
        bx.clicked.connect(self.accept)
        h.addWidget(bx)
        layout.addLayout(h)

        # Section 1: Thông tin cá nhân
        layout.addWidget(self._section_label("THÔNG TIN CÁ NHÂN"))

        for row_items in [
            [("HỌ VÀ TÊN", self.guest.full_name),
             ("MÃ KT", f"KT{self.guest.id:03d}" if self.guest.id else "—")],
            [("SỐ ĐIỆN THOẠI", str(self.guest.phone)),
             ("CCCD/CMND", str(self.guest.id_card))],
            [("EMAIL", self.guest.email),
             ("NGHỀ NGHIỆP", self.guest.occupation)],
        ]:
            r = QHBoxLayout()
            r.setSpacing(10)
            for lb, vl in row_items:
                r.addWidget(self._box(lb, vl))
            layout.addLayout(r)
        layout.addWidget(self._box("ĐỊA CHỈ THƯỜNG TRÚ", self.guest.address))

        # Section 2: Thông tin hợp đồng
        layout.addWidget(self._section_label("THÔNG TIN HỢP ĐỒNG"))

        st = self.contract_info.get('status_text', '—')
        sc_map = {'● Đang thuê': '#38a169', '● Sắp hết hạn': '#dd6b20', '● Đã chấm dứt': '#e53e3e'}
        sc = sc_map.get(st, '#718096')
        clean_st = st.replace('● ', '') if st.startswith('●') else st

        r5 = QHBoxLayout()
        r5.setSpacing(10)
        r5.addWidget(self._box("PHÒNG", self.contract_info.get('room_number', '—')))
        r5.addWidget(self._box("TRẠNG THÁI", clean_st, color=sc))
        layout.addLayout(r5)

        r6 = QHBoxLayout()
        r6.setSpacing(10)
        r6.addWidget(self._box("NGÀY BẮT ĐẦU", self.contract_info.get('start_date', '—')))
        r6.addWidget(self._box("NGÀY KẾT THÚC", self.contract_info.get('end_date', '—')))
        layout.addLayout(r6)

        rent = self.contract_info.get('monthly_rent')
        dep = self.contract_info.get('deposit')
        r7 = QHBoxLayout()
        r7.setSpacing(10)
        r7.addWidget(self._box("GIÁ THUÊ", f"{int(rent):,.0f}đ/tháng" if rent else "—"))
        r7.addWidget(self._box("TIỀN CỌC", f"{int(dep):,.0f}đ" if dep else "—"))
        layout.addLayout(r7)

        layout.addStretch()

        # Buttons
        br = QHBoxLayout()
        br.setSpacing(10)
        btn_contract = QPushButton("📋 Xem hợp đồng")
        btn_contract.setStyleSheet(
            "QPushButton{background:#ebf8ff;color:#3182ce;border:1px solid #bee3f8;"
            "border-radius:8px;padding:10px 16px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background:#bee3f8;}")
        btn_contract.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_contract.clicked.connect(self._on_view_contract)
        br.addWidget(btn_contract)
        layout.addLayout(br)

        btn_terminate = QPushButton("⛔ Chấm dứt hợp đồng")
        btn_terminate.setStyleSheet(
            "QPushButton{background:white;color:#e53e3e;border:1px solid #fed7d7;"
            "border-radius:8px;padding:10px 16px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background:#fff5f5;}")
        btn_terminate.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_terminate.clicked.connect(self._on_terminate)
        layout.addWidget(btn_terminate)

        scroll.setWidget(container)
        outer.addWidget(scroll)

    def _on_terminate(self):
        if ask_danger(self, "Xác nhận chấm dứt hợp đồng",
            f"Bạn có chắc muốn chấm dứt hợp đồng của khách {self.guest.full_name}?\n"
            f"Phòng sẽ được giải phóng, tài khoản khách thuê vẫn được giữ lại.",
            yes_text="Chấm dứt", no_text="Hủy bỏ"):
            self.terminate_requested.emit(self.guest.id)
            self.accept()

    def _on_view_contract(self):
        """Mở dialog xem hợp đồng."""
        from PyQt6.QtWidgets import QTextBrowser, QFileDialog
        from datetime import datetime

        rent = self.contract_info.get('monthly_rent', 0)
        dep = self.contract_info.get('deposit', 0)

        contract_data = {
            'contract_number': self.contract_info.get('contract_number', '—'),
            'sign_date': datetime.now().strftime('%d/%m/%Y'),
            'landlord_name': 'Tạ Nhật Anh',
            'landlord_phone': '0364216007',
            'landlord_address': '2 đường số 10, P. Tam Bình, Thủ Đức, TP.HCM',
            'guest_name': self.guest.full_name,
            'guest_dob': getattr(self.guest, 'dob', '—') or '—',
            'guest_gender': getattr(self.guest, 'gender', '—') or '—',
            'guest_cccd': str(self.guest.id_card) if self.guest.id_card else '—',
            'guest_phone': str(self.guest.phone) if self.guest.phone else '—',
            'guest_email': getattr(self.guest, 'email', '—') or '—',
            'room_number': self.contract_info.get('room_number', '—'),
            'room_floor': '—',
            'room_area': '—',
            'monthly_rent': f'{int(rent):,}' if rent else '—',
            'deposit': f'{int(dep):,}' if dep else '—',
            'start_date': self.contract_info.get('start_date', '—'),
            'end_date': self.contract_info.get('end_date', '—'),
        }

        from utils.pdf_generator import PDFGenerator
        html = PDFGenerator._build_contract_html(contract_data)

        dlg = QDialog(self)
        dlg.setWindowTitle(f"Hợp đồng — {self.guest.full_name}")
        dlg.setFixedSize(650, 700)
        dlg.setStyleSheet("QDialog { background-color: white; }")

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        browser = QTextBrowser()
        browser.setHtml(html)
        browser.setStyleSheet("QTextBrowser { border: none; padding: 10px; }")
        lay.addWidget(browser)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(20, 10, 20, 15)
        btn_row.setSpacing(10)
        btn_row.addStretch()

        def _export():
            default_name = f"HopDong_{self.guest.full_name}.pdf"
            file_path, _ = QFileDialog.getSaveFileName(
                dlg, "Lưu hợp đồng", default_name,
                "PDF Files (*.pdf);;All Files (*)")
            if not file_path:
                return
            try:
                result = PDFGenerator.export_contract_pdf(contract_data, file_path)
                show_success(dlg, "Xuất thành công", f"Đã lưu tại:\n{result}")
                import os
                os.startfile(result)
            except Exception as e:
                show_error(dlg, "Lỗi", f"Không thể xuất PDF:\n{e}")

        btn_export = QPushButton("📄 Xuất PDF")
        btn_export.setStyleSheet(
            "QPushButton { background-color: #0b8480; color: white; border: none; "
            "border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background-color: #096c69; }")
        btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export.clicked.connect(_export)
        btn_row.addWidget(btn_export)

        btn_close = QPushButton("Đóng")
        btn_close.setStyleSheet(
            "QPushButton { background: #f7fafc; color: #4a5568; border: 1px solid #cbd5e0; "
            "border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background: #edf2f7; }")
        btn_close.clicked.connect(dlg.close)
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

        dlg.exec()


# ── Main View ─────────────────────────────────────────────
class GuestManagementView(QWidget):
    """Widget quản lý khách thuê — Được nhúng vào AdminWindow."""

    HEADERS = ["Mã KT", "Khách thuê", "Phòng", "CCCD/CMND", "Ngày bắt đầu",
               "Ngày kết thúc", "Trạng thái HĐ", "Thao tác"]

    STATUS_COLORS = {
        'active': ('● Đang thuê', '#3182ce'),
        'expired': ('● Sắp hết hạn', '#dd6b20'),
        'terminated': ('● Đã chấm dứt', '#e53e3e'),
    }

    def __init__(self, guest_service=None, contract_service=None, room_service=None, parent=None):
        super().__init__(parent)
        self.guest_service = guest_service
        self.contract_service = contract_service
        self.room_service = room_service
        self._build_ui()
        self.refresh_data()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 15, 20, 15)
        main_layout.setSpacing(12)

        # ── Stat cards row ──
        stats_row = QHBoxLayout()
        stats_row.setSpacing(15)
        self.card_total = StatCard("👥", "0", "Tổng khách thuê", "#ebf8ff")
        self.card_active = StatCard("✅", "0", "Đang thuê", "#e6fffa")
        self.card_expiring = StatCard("📋", "0", "Sắp hết hạn", "#fffaf0")
        self.card_terminated = StatCard("🚫", "0", "Đã chấm dứt", "#fff5f5")
        for card in [self.card_total, self.card_active, self.card_expiring, self.card_terminated]:
            stats_row.addWidget(card)
        main_layout.addLayout(stats_row)

        # ── Search + filter bar ──
        bar = QHBoxLayout()
        bar.setSpacing(10)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Tìm kiếm theo tên, SĐT, phòng...")
        self.search_input.setStyleSheet(
            "QLineEdit { background-color: #f7fafc; border: 1px solid #e2e8f0; "
            "border-radius: 8px; padding: 8px 15px; font-size: 13px; }")
        self.search_input.textChanged.connect(self._on_search)
        bar.addWidget(self.search_input, 1)

        self.filter_status = QComboBox()
        self.filter_status.addItems(["Tất cả trạng thái", "Đang thuê", "Sắp hết hạn", "Đã chấm dứt"])
        self.filter_status.setStyleSheet(
            "QComboBox { background-color: #f7fafc; border: 1px solid #e2e8f0; "
            "border-radius: 8px; padding: 8px 12px; font-size: 13px; }")
        self.filter_status.currentIndexChanged.connect(lambda: self.refresh_data())
        bar.addWidget(self.filter_status)

        bar.addStretch()

        btn_export_excel = QPushButton("📊 Xuất Excel")
        btn_export_excel.setStyleSheet(
            "QPushButton { background-color: #2b6cb0; color: white; border: none; "
            "border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background-color: #2c5282; }")
        btn_export_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export_excel.clicked.connect(self._on_export_excel)
        bar.addWidget(btn_export_excel)



        main_layout.addLayout(bar)

        # ── Table ──
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.HEADERS))
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 60)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(7, 150)
        for col in range(1, 7):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white; border: 1px solid #e2e8f0; border-radius: 8px;
                gridline-color: #edf2f7; font-size: 12px;
            }
            QTableWidget::item { padding: 4px; }
            QTableWidget::item:alternate { background-color: #f7fafc; }
            QHeaderView::section {
                background-color: #f7fafc; color: #718096; font-weight: bold;
                border: none; padding: 8px 4px; font-size: 11px;
            }
        """)
        main_layout.addWidget(self.table, 1)

        self.lbl_count = QLabel("Hiển thị 0 khách thuê")
        self.lbl_count.setStyleSheet("color: #a0aec0; font-size: 12px;")
        main_layout.addWidget(self.lbl_count)

    # ── Data ──
    def refresh_data(self):
        if not self.guest_service:
            return

        guests = self.guest_service.get_all_guests()

        contracts_by_guest = {}
        rooms_by_id = {}
        if self.contract_service:
            try:
                for c in self.contract_service.get_all():
                    contracts_by_guest[c.guest_id] = c
            except Exception:
                pass
        if self.room_service:
            try:
                for r in self.room_service.get_all_rooms():
                    rooms_by_id[r.id] = r
            except Exception:
                pass

        rows = []
        for g in guests:
            contract = contracts_by_guest.get(g.id)
            room = rooms_by_id.get(contract.room_id) if contract else None
            status = contract.status if contract else 'terminated'
            rows.append({
                'guest': g, 'contract': contract,
                'room_number': room.room_number if room else '—',
                'start_date': contract.start_date if contract else '—',
                'end_date': contract.end_date if contract else '—',
                'status': status,
            })

        search_q = self.search_input.text().strip().lower()
        if search_q:
            rows = [r for r in rows if (
                search_q in r['guest'].full_name.lower() or
                search_q in str(r['guest'].phone).lower() or
                search_q in r['room_number'].lower()
            )]

        filter_map = {1: 'active', 2: 'expired', 3: 'terminated'}
        status_filter = filter_map.get(self.filter_status.currentIndex())
        if status_filter:
            rows = [r for r in rows if r['status'] == status_filter]

        # Stats
        all_guests = self.guest_service.get_all_guests()
        self.card_total.val_lbl.setText(str(len(all_guests)))
        active_count = sum(1 for g in all_guests if contracts_by_guest.get(g.id) and contracts_by_guest[g.id].status == 'active')
        expired_count = sum(1 for g in all_guests if contracts_by_guest.get(g.id) and contracts_by_guest[g.id].status == 'expired')
        terminated_count = sum(1 for g in all_guests if contracts_by_guest.get(g.id) and contracts_by_guest[g.id].status == 'terminated')
        self.card_active.val_lbl.setText(str(active_count))
        self.card_expiring.val_lbl.setText(str(expired_count))
        self.card_terminated.val_lbl.setText(str(terminated_count))

        # Table
        self.table.setRowCount(len(rows))
        for i, row_data in enumerate(rows):
            g = row_data['guest']
            self.table.setRowHeight(i, 55)

            self.table.setItem(i, 0, QTableWidgetItem(f"KT{g.id:03d}" if g.id else "—"))

            # Name + phone widget
            initials = "".join([w[0].upper() for w in g.full_name.split()[:2]]) if g.full_name else "?"
            name_widget = QWidget()
            name_widget.setStyleSheet("background: transparent;")
            name_layout = QHBoxLayout(name_widget)
            name_layout.setContentsMargins(4, 4, 4, 4)
            name_layout.setSpacing(8)

            avatar = QLabel(initials)
            avatar.setFixedSize(32, 32)
            avatar.setAlignment(Qt.AlignmentFlag.AlignCenter)
            colors = ['#0b8480', '#3182ce', '#dd6b20', '#e53e3e', '#805ad5']
            avatar.setStyleSheet(f"background-color: {colors[g.id % len(colors) if g.id else 0]}; "
                                  "color: white; border-radius: 16px; font-weight: bold; font-size: 12px;")
            name_layout.addWidget(avatar)

            info = QVBoxLayout()
            info.setSpacing(0)
            lbl_n = QLabel(g.full_name)
            lbl_n.setStyleSheet("font-weight: bold; color: #2d3748; font-size: 12px; background: transparent;")
            lbl_n.setWordWrap(True)
            info.addWidget(lbl_n)
            lbl_p = QLabel(str(g.phone))
            lbl_p.setStyleSheet("color: #a0aec0; font-size: 10px; background: transparent;")
            info.addWidget(lbl_p)
            name_layout.addLayout(info)
            name_layout.addStretch()
            self.table.setCellWidget(i, 1, name_widget)

            self.table.setItem(i, 2, QTableWidgetItem(row_data['room_number']))
            self.table.setItem(i, 3, QTableWidgetItem(str(g.id_card) if g.id_card else '—'))
            self.table.setItem(i, 4, QTableWidgetItem(str(row_data['start_date']) if row_data['start_date'] else '—'))
            self.table.setItem(i, 5, QTableWidgetItem(str(row_data['end_date']) if row_data['end_date'] else '—'))

            status_text, status_color = self.STATUS_COLORS.get(row_data['status'], ('—', '#718096'))
            status_lbl = QLabel(status_text)
            status_lbl.setStyleSheet(f"color: {status_color}; font-weight: bold; font-size: 11px; background: transparent;")
            status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setCellWidget(i, 6, status_lbl)

            # Action button
            actions = QWidget()
            actions.setStyleSheet("background: transparent;")
            actions_layout = QHBoxLayout(actions)
            actions_layout.setContentsMargins(4, 4, 4, 4)
            actions_layout.setSpacing(4)
            for text, color, bg, callback in [
                ("Chi tiết", "#3182ce", "#ebf8ff", lambda _, g=g, rd=row_data: self._on_detail(g, rd)),
                ("Xuất HĐ", "#0b8480", "#e6fffa", lambda _, g=g, rd=row_data: self._on_export_contract(g, rd)),
            ]:
                btn = QPushButton(text)
                btn.setStyleSheet(
                    f"QPushButton {{ background-color: {bg}; color: {color}; border: none; "
                    f"border-radius: 4px; padding: 4px 8px; font-weight: bold; font-size: 10px; }}"
                    f"QPushButton:hover {{ opacity: 0.8; }}")
                btn.setCursor(Qt.CursorShape.PointingHandCursor)
                btn.setFixedHeight(24)
                btn.clicked.connect(callback)
                actions_layout.addWidget(btn)
            self.table.setCellWidget(i, 7, actions)

        self.lbl_count.setText(f"Hiển thị {len(rows)} khách thuê")

    def _on_search(self, text):
        self.refresh_data()

    # ── CRUD ──
    def _on_edit(self, guest: Guest):
        rooms = self.room_service.get_all_rooms() if self.room_service else []
        dlg = GuestFormDialog(guest=guest, rooms=rooms, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.result_data
            guest.full_name = data['full_name']
            guest.phone = data['phone']
            guest.id_card = data['id_card']
            guest.address = data['address']
            ok, msg = self.guest_service.update_guest(guest)
            if ok:
                show_success(self, "Thành công", msg)
                self.refresh_data()
            else:
                show_warning(self, "Lỗi", msg)

    def _on_detail(self, guest: Guest, row_data: dict):
        contract = row_data.get('contract')
        contract_info = {
            'contract_number': contract.contract_number if contract else '—',
            'room_number': row_data['room_number'],
            'start_date': row_data['start_date'],
            'end_date': row_data['end_date'],
            'status_text': self.STATUS_COLORS.get(row_data['status'], ('—', '#718096'))[0],
            'monthly_rent': contract.monthly_rent if contract else None,
            'deposit': contract.deposit if contract else None,
        }
        dlg = GuestDetailDialog(guest, contract_info=contract_info, parent=self)
        dlg.terminate_requested.connect(lambda gid: self._on_terminate_confirmed(gid, row_data))
        dlg.exec()

    def _on_terminate_confirmed(self, guest_id: int, row_data: dict):
        contract = row_data.get('contract')
        if not contract:
            show_warning(self, "Lỗi", "Khách này hiện không có hợp đồng để chấm dứt.")
            return
        if contract.status == 'terminated':
            show_warning(self, "Lỗi", "Hợp đồng này đã được chấm dứt trước đó.")
            return
        ok, msg = self.contract_service.terminate_contract(contract.id)
        if ok:
            show_success(self, "Thành công", msg)
            self.refresh_data()
        else:
            show_warning(self, "Lỗi", msg)

    def _on_export_contract(self, guest: Guest, row_data: dict):
        """Xuất hợp đồng PDF cho khách thuê."""
        from PyQt6.QtWidgets import QFileDialog
        from datetime import datetime

        contract = row_data.get('contract')
        if not contract:
            show_warning(self, "Lỗi", "Khách này chưa có hợp đồng.")
            return

        room = None
        if self.room_service:
            room = self.room_service.get_room_by_id(contract.room_id)

        price = contract.monthly_rent or (room.price if room else 0)
        deposit = contract.deposit or (room.deposit if room else 0)

        contract_data = {
            'contract_number': contract.contract_number or f'HD{contract.id}',
            'sign_date': datetime.now().strftime('%d/%m/%Y'),
            'landlord_name': 'Tạ Nhật Anh',
            'landlord_phone': '0364216007',
            'landlord_address': '2 đường số 10, P. Tam Bình, Thủ Đức, TP.HCM',
            'guest_name': guest.full_name,
            'guest_dob': getattr(guest, 'dob', '—') or '—',
            'guest_gender': getattr(guest, 'gender', '—') or '—',
            'guest_cccd': str(guest.id_card) if guest.id_card else '—',
            'guest_phone': str(guest.phone) if guest.phone else '—',
            'guest_email': getattr(guest, 'email', '—') or '—',
            'room_number': room.room_number if room else '—',
            'room_floor': room.floor if room else '—',
            'room_area': room.area if room else '—',
            'monthly_rent': f'{int(price):,}',
            'deposit': f'{int(deposit):,}',
            'start_date': contract.start_date or '—',
            'end_date': contract.end_date or '—',
        }

        default_name = f"HopDong_{contract_data['contract_number']}_{guest.full_name}.pdf"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu hợp đồng", default_name,
            "PDF Files (*.pdf);;All Files (*)")
        if not file_path:
            return

        try:
            from utils.pdf_generator import PDFGenerator
            result = PDFGenerator.export_contract_pdf(contract_data, file_path)
            show_success(self, "Xuất thành công",
                         f"Hợp đồng đã được lưu tại:\n{result}")
            import os as _os
            _os.startfile(result)
        except Exception as e:
            show_error(self, "Lỗi", f"Không thể xuất PDF:\n{e}")

    def _on_export_excel(self):
        """Xuất danh sách khách thuê ra Excel (tất cả)."""
        from PyQt6.QtWidgets import QFileDialog
        from datetime import datetime

        default_name = f"DanhSachKhachThue_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Lưu danh sách khách thuê", default_name,
            "Excel Files (*.xlsx);;All Files (*)")
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Danh sách khách thuê"

            ws.merge_cells('A1:I1')
            t = ws.cell(row=1, column=1, value="DANH SÁCH KHÁCH THUÊ")
            t.font = Font(name='Arial', bold=True, size=13, color='0B8480')
            t.alignment = Alignment(horizontal='center')

            hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            hfill = PatternFill(start_color='0B8480', end_color='0B8480', fill_type='solid')
            ha = Alignment(horizontal='center', vertical='center')
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            for col, h in enumerate(['STT', 'Họ tên', 'SĐT', 'CCCD/CMND', 'Email',
                                     'Phòng', 'Ngày bắt đầu', 'Ngày kết thúc', 'Trạng thái HĐ'], 1):
                c = ws.cell(row=3, column=col, value=h)
                c.font = hf; c.fill = hfill; c.alignment = ha; c.border = border

            guests = self.guest_service.get_all_guests() if self.guest_service else []
            contracts_by_guest = {}
            rooms_by_id = {}
            if self.contract_service:
                for c in self.contract_service.get_all():
                    contracts_by_guest[c.guest_id] = c
            if self.room_service:
                for r in self.room_service.get_all_rooms():
                    rooms_by_id[r.id] = r

            sm = {'active': 'Đang thuê', 'expired': 'Sắp hết hạn', 'terminated': 'Đã chấm dứt'}
            df = Font(name='Arial', size=10)
            da = Alignment(vertical='center')

            for i, g in enumerate(guests, 1):
                ct = contracts_by_guest.get(g.id)
                room = rooms_by_id.get(ct.room_id) if ct else None
                vals = [i, g.full_name, str(g.phone) if g.phone else '—',
                        str(g.id_card) if g.id_card else '—',
                        getattr(g, 'email', '—') or '—',
                        room.room_number if room else '—',
                        ct.start_date if ct else '—', ct.end_date if ct else '—',
                        sm.get(ct.status, ct.status) if ct else '—']
                for col, val in enumerate(vals, 1):
                    c = ws.cell(row=i + 3, column=col, value=val)
                    c.font = df; c.alignment = da; c.border = border

            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                ml = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        ml = max(ml, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(ml + 4, 30)

            wb.save(file_path)
            show_success(self, "Xuất thành công", f"Đã xuất {len(guests)} khách thuê.\nLưu tại: {file_path}")
            import os
            os.startfile(file_path)
        except Exception as e:
            show_error(self, "Lỗi", f"Không thể xuất Excel:\n{e}")

