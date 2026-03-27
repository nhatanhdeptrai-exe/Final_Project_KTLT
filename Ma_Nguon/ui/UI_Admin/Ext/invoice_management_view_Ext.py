
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QLineEdit, QComboBox, QMessageBox, QDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView,
    QSizePolicy, QScrollArea, QDateEdit
)
from PyQt6.QtCore import Qt, pyqtSignal, QDate
from PyQt6.QtGui import QFont, QColor

from models.invoice import Invoice
from ui.UI_Admin.generated.ui_dialog_them_hoa_don_UI import Ui_DialogFormInvoice
from ui.UI_Common.custom_popup import show_success, show_error, show_warning, show_info, ask_question, ask_danger



_inv_stat_counter = 0

class InvoiceStatCard(QFrame):
    def __init__(self, icon: str, value: str, label: str, bg_color: str, val_color: str = "#2d3748", parent=None):
        super().__init__(parent)
        global _inv_stat_counter
        _inv_stat_counter += 1
        obj_name = f"invStatCard_{_inv_stat_counter}"
        self.setObjectName(obj_name)
        self.setStyleSheet(
            f"QFrame#{obj_name} {{ background-color: white; border: 1px solid #e2e8f0;"
            f" border-radius: 10px; padding: 12px; }}"
            f" QFrame#{obj_name} QLabel {{ border: none; }}")
        layout = QHBoxLayout(self)
        layout.setSpacing(10)
        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet(f"font-size: 24px; background: {bg_color}; border-radius: 8px; padding: 8px; border: none;")
        icon_lbl.setFixedSize(44, 44)
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_lbl)
        info = QVBoxLayout()
        info.setSpacing(2)
        self.val_lbl = QLabel(value)
        self.val_lbl.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        self.val_lbl.setStyleSheet(f"color: {val_color}; border: none;")
        info.addWidget(self.val_lbl)
        desc_lbl = QLabel(label)
        desc_lbl.setStyleSheet("color: #718096; font-size: 11px; border: none;")
        info.addWidget(desc_lbl)
        layout.addLayout(info)


# ── Invoice Form Dialog ──────────────────────────────────
class InvoiceFormDialog(QDialog):
    """Dialog tạo/cập nhật hóa đơn."""

    def __init__(self, guests_rooms=None, contracts=None, invoice: Invoice = None, parent=None):
        super().__init__(parent)
        self.invoice = invoice
        self.ui = Ui_DialogFormInvoice()
        self.ui.setupUi(self)
        self.contracts = contracts or []
        self.guests_rooms = guests_rooms or {}  # {contract_id: (guest_name, room_number)}

        # Replace due date with QDateEdit
        self._replace_date_edit('inpDueDate')

        # Set current month
        today = QDate.currentDate()
        self.ui.inpMonth.setText(f"{today.month():02d}/{today.year()}")
        self.ui.inpMonth.setReadOnly(True)

        # Populate guest/contract combo
        self.ui.cbGuest.clear()
        self.ui.cbGuest.addItem("— Chọn khách thuê —")
        for c in self.contracts:
            info = self.guests_rooms.get(c.id, (f"KT{c.guest_id:03d}", f"P{c.room_id}"))
            self.ui.cbGuest.addItem(f"{info[0]} — {info[1]}", c.id)

        # Auto-calculate electricity/water from readings, then total
        self.ui.cbGuest.currentIndexChanged.connect(self._on_guest_changed)
        for field in ['inpElecOld', 'inpElecNew', 'inpWaterOld', 'inpWaterNew',
                       'inpOtherCost', 'inpRoomPrice']:
            widget = getattr(self.ui, field, None)
            if widget:
                widget.textChanged.connect(self._calc_total)

        if invoice:
            self.setWindowTitle("📝 Cập nhật hóa đơn")
            self._fill_form(invoice)
        else:
            self.setWindowTitle("➕ Tạo hóa đơn mới")

        self.ui.btnCancel.clicked.connect(self.reject)
        self.ui.btnSave.clicked.connect(self._on_save)

    def _replace_date_edit(self, attr_name: str):
        old = getattr(self.ui, attr_name, None)
        if not old:
            return
        de = QDateEdit(old.parentWidget())
        de.setCalendarPopup(True)
        de.setDisplayFormat("dd/MM/yyyy")
        de.setDate(QDate.currentDate().addDays(5))
        de.setStyleSheet(
            "QDateEdit { background-color: white; border: 1px solid #e2e8f0; "
            "border-radius: 6px; padding: 8px; font-size: 13px; color: #1a202c; }"
            "QDateEdit:focus { border: 1px solid #3182ce; }"
            "QDateEdit::drop-down { border: none; padding-right: 8px; }")
        de.setMinimumHeight(old.minimumHeight() if old.minimumHeight() > 0 else 36)
        form_grid = self.ui.formGrid
        for i in range(form_grid.count()):
            item = form_grid.itemAt(i)
            if item and item.layout():
                sub = item.layout()
                if sub.indexOf(old) >= 0:
                    sub.replaceWidget(old, de)
                    old.hide(); old.deleteLater()
                    setattr(self.ui, attr_name, de)
                    return

    def _on_guest_changed(self, index):
        if index <= 0:
            return
        contract_id = self.ui.cbGuest.currentData()
        for c in self.contracts:
            if c.id == contract_id:
                self.ui.inpRoomPrice.setText(str(c.monthly_rent))
                break

    def _calc_total(self):
        try:
            # Electricity: (new - old) * rate
            ELEC_RATE = 3500  # VND/kWh
            WATER_RATE = 25000  # VND/m³
            elec_old = int(self.ui.inpElecOld.text() or '0')
            elec_new = int(self.ui.inpElecNew.text() or '0')
            elec_cost = max(0, elec_new - elec_old) * ELEC_RATE
            self.ui.inpElecCost.setText(f"{elec_cost:,}".replace(',', '.'))

            water_old = int(self.ui.inpWaterOld.text() or '0')
            water_new = int(self.ui.inpWaterNew.text() or '0')
            water_cost = max(0, water_new - water_old) * WATER_RATE
            self.ui.inpWaterCost.setText(f"{water_cost:,}".replace(',', '.'))

            rent = int(self.ui.inpRoomPrice.text().replace('.', '').replace(',', '') or '0')
            other = int(self.ui.inpOtherCost.text().replace('.', '').replace(',', '') or '0')
            total = rent + elec_cost + water_cost + other
            self.ui.inpTotalCost.setText(f"{total:,}".replace(',', '.'))
        except ValueError:
            pass

    def _fill_form(self, inv: Invoice):
        self.ui.inpMonth.setText(f"{inv.month:02d}/{inv.year}")
        self.ui.inpRoomPrice.setText(str(inv.room_rent))
        self.ui.inpElecCost.setText(str(inv.electricity_cost))
        self.ui.inpWaterCost.setText(str(inv.water_cost))
        self.ui.inpOtherCost.setText(str(inv.other_fees))
        self.ui.inpTotalCost.setText(str(inv.total_amount))
        # Select contract
        for i in range(self.ui.cbGuest.count()):
            if self.ui.cbGuest.itemData(i) == inv.contract_id:
                self.ui.cbGuest.setCurrentIndex(i)
                break

    def _on_save(self):
        if self.ui.cbGuest.currentIndex() == 0:
            show_warning(self, "Lỗi", "Vui lòng chọn khách thuê"); return
        if not self.ui.inpRoomPrice.text().strip():
            show_warning(self, "Lỗi", "Vui lòng nhập tiền phòng"); return

        today = QDate.currentDate()
        rent = int(self.ui.inpRoomPrice.text().replace('.', '').replace(',', '') or '0')
        elec = int(self.ui.inpElecCost.text().replace('.', '').replace(',', '') or '0')
        water = int(self.ui.inpWaterCost.text().replace('.', '').replace(',', '') or '0')
        other = int(self.ui.inpOtherCost.text().replace('.', '').replace(',', '') or '0')
        due = self.ui.inpDueDate.date().toString("yyyy-MM-dd")

        self.result_data = {
            'contract_id': self.ui.cbGuest.currentData(),
            'month': today.month(),
            'year': today.year(),
            'room_rent': rent,
            'electricity_cost': elec,
            'water_cost': water,
            'other_fees': other,
            'total_amount': rent + elec + water + other,
            'due_date': due,
        }
        self.accept()


# ── Invoice Detail Dialog ────────────────────────────────
class InvoiceDetailDialog(QDialog):
    """Dialog xem chi tiết hóa đơn."""
    delete_requested = pyqtSignal(int)

    def __init__(self, invoice: Invoice, extra: dict = None, parent=None):
        super().__init__(parent)
        self.invoice = invoice
        self.extra = extra or {}
        self.setWindowTitle(f"Chi tiết — {invoice.invoice_number}")
        self.setFixedSize(450, 550)
        self.setStyleSheet("QDialog { background-color: white; } QLabel { color: #2d3748; }")
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowType.WindowContextHelpButtonHint)
        self._build_ui()

    def _box(self, label: str, value: str, color: str = None) -> QFrame:
        box = QFrame()
        box.setStyleSheet("QFrame{background:#f7fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px;}")
        lay = QVBoxLayout(box)
        lay.setContentsMargins(0, 0, 0, 0); lay.setSpacing(3)
        lbl = QLabel(label)
        lbl.setStyleSheet("color:#a0aec0;font-size:10px;font-weight:bold;border:none;background:transparent;")
        lay.addWidget(lbl)
        if color:
            val = QLabel(f"● {value}")
            val.setStyleSheet(f"color:{color};font-size:13px;font-weight:bold;border:none;background:transparent;")
        else:
            val = QLabel(str(value) if value else "—")
            val.setStyleSheet("color:#2d3748;font-size:13px;font-weight:bold;border:none;background:transparent;")
        val.setWordWrap(True)
        lay.addWidget(val)
        return box

    def _section(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet("color:#718096;font-size:11px;font-weight:bold;letter-spacing:1px;")
        return lbl

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:white;}")
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setSpacing(8)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header
        h = QHBoxLayout()
        t = QLabel("🧾 Chi tiết hóa đơn")
        t.setFont(QFont("Segoe UI", 15, QFont.Weight.Bold))
        h.addWidget(t); h.addStretch()
        bx = QPushButton("✕"); bx.setFixedSize(28, 28)
        bx.setStyleSheet("QPushButton{background:transparent;border:none;font-size:16px;color:#a0aec0;}"
                          "QPushButton:hover{color:#e53e3e;}")
        bx.clicked.connect(self.accept); h.addWidget(bx)
        layout.addLayout(h)

        inv = self.invoice
        status_map = {'paid': ('Đã thanh toán', '#38a169'), 'unpaid': ('Chưa thanh toán', '#e53e3e'), 'overdue': ('Quá hạn', '#dd6b20')}
        st_text, st_color = status_map.get(inv.status, ('—', '#718096'))

        layout.addWidget(self._section("THÔNG TIN CHUNG"))
        for row in [
            [("MÃ HĐ", inv.invoice_number), ("THÁNG", f"{inv.month:02d}/{inv.year}")],
            [("KHÁCH THUÊ", self.extra.get('guest_name', '—')), ("PHÒNG", self.extra.get('room_number', '—'))],
            [("TRẠNG THÁI", st_text)],
        ]:
            r = QHBoxLayout(); r.setSpacing(8)
            for lb, vl in row:
                color = st_color if lb == "TRẠNG THÁI" else None
                r.addWidget(self._box(lb, vl, color=color))
            layout.addLayout(r)

        layout.addWidget(self._section("CHI TIẾT CHI PHÍ"))
        for row in [
            [("TIỀN PHÒNG", f"{inv.room_rent:,.0f}đ"), ("TIỀN ĐIỆN", f"{inv.electricity_cost:,.0f}đ")],
            [("TIỀN NƯỚC", f"{inv.water_cost:,.0f}đ"), ("PHÍ KHÁC", f"{inv.other_fees:,.0f}đ")],
        ]:
            r = QHBoxLayout(); r.setSpacing(8)
            for lb, vl in row:
                r.addWidget(self._box(lb, vl))
            layout.addLayout(r)

        # Total
        total_box = QFrame()
        total_box.setStyleSheet("QFrame{background:#ebf8ff;border:1px solid #bee3f8;border-radius:8px;padding:12px;}")
        tl = QHBoxLayout(total_box)
        tl.setContentsMargins(0, 0, 0, 0)
        tl_lbl = QLabel("TỔNG CỘNG")
        tl_lbl.setStyleSheet("color:#3182ce;font-size:12px;font-weight:bold;border:none;background:transparent;")
        tl.addWidget(tl_lbl)
        tl.addStretch()
        tl_val = QLabel(f"{inv.total_amount:,.0f}đ")
        tl_val.setStyleSheet("color:#2d3748;font-size:18px;font-weight:bold;border:none;background:transparent;")
        tl.addWidget(tl_val)
        layout.addWidget(total_box)



        layout.addStretch()

        # Buttons
        br = QHBoxLayout(); br.setSpacing(10)
        btn_del = QPushButton("🗑️ Xóa hóa đơn")
        btn_del.setStyleSheet(
            "QPushButton{background:white;color:#e53e3e;border:1px solid #fed7d7;"
            "border-radius:8px;padding:10px 16px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background:#fff5f5;}")
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.clicked.connect(self._on_delete)
        br.addWidget(btn_del)
        layout.addLayout(br)

        scroll.setWidget(container)
        outer.addWidget(scroll)

    def _on_delete(self):
        if ask_danger(self, "Xác nhận xóa",
            f"Xóa hóa đơn {self.invoice.invoice_number}?",
            yes_text="Xóa", no_text="Hủy bỏ"):
            self.delete_requested.emit(self.invoice.id)
            self.accept()


# ── Main View ────────────────────────────────────────────
class InvoiceManagementView(QWidget):
    """Widget quản lý hóa đơn — Được nhúng vào AdminWindow."""

    HEADERS = ["Mã HĐ", "Khách thuê", "Phòng", "Tháng", "Tiền phòng",
               "Tiền điện", "Tiền nước", "Tổng tiền", "Trạng thái", "Thao tác"]

    STATUS_COLORS = {
        'paid': ('● Đã thanh toán', '#38a169'),
        'unpaid': ('● Chưa thanh toán', '#e53e3e'),
        'overdue': ('● Quá hạn', '#dd6b20'),
    }

    def __init__(self, invoice_service=None, contract_service=None,
                 guest_service=None, room_service=None, parent=None):
        super().__init__(parent)
        self.invoice_service = invoice_service
        self.contract_service = contract_service
        self.guest_service = guest_service
        self.room_service = room_service
        self._build_ui()
        self.refresh_data()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 15, 20, 15)
        main_layout.setSpacing(12)

        # ── Stat cards ──
        stats_row = QHBoxLayout()
        stats_row.setSpacing(15)
        self.card_total = InvoiceStatCard("🧾", "0", "Tổng hóa đơn", "#ebf8ff")
        self.card_paid = InvoiceStatCard("✅", "0", "Đã thanh toán", "#e6fffa", "#38a169")
        self.card_unpaid = InvoiceStatCard("⏳", "0", "Chưa thanh toán", "#fff5f5", "#e53e3e")
        self.card_revenue = InvoiceStatCard("💰", "0", "Tổng thu tháng này", "#fffaf0", "#dd6b20")
        for card in [self.card_total, self.card_paid, self.card_unpaid, self.card_revenue]:
            stats_row.addWidget(card)
        main_layout.addLayout(stats_row)

        # ── Search + filter bar ──
        bar = QHBoxLayout()
        bar.setSpacing(10)
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Tìm tên, phòng, mã HĐ...")
        self.search_input.setStyleSheet(
            "QLineEdit { background-color: #f7fafc; border: 1px solid #e2e8f0; "
            "border-radius: 8px; padding: 8px 15px; font-size: 13px; }")
        self.search_input.textChanged.connect(self._on_search)
        bar.addWidget(self.search_input, 1)

        self.filter_month = QComboBox()
        self.filter_month.addItem("Tất cả tháng")
        self.filter_month.setStyleSheet(
            "QComboBox { background-color: #f7fafc; border: 1px solid #e2e8f0; "
            "border-radius: 8px; padding: 8px 12px; font-size: 13px; }")
        self.filter_month.currentIndexChanged.connect(lambda: self.refresh_data())
        bar.addWidget(self.filter_month)

        self.filter_status = QComboBox()
        self.filter_status.addItems(["Tất cả trạng thái", "Đã thanh toán", "Chưa thanh toán", "Quá hạn"])
        self.filter_status.setStyleSheet(
            "QComboBox { background-color: #f7fafc; border: 1px solid #e2e8f0; "
            "border-radius: 8px; padding: 8px 12px; font-size: 13px; }")
        self.filter_status.currentIndexChanged.connect(lambda: self.refresh_data())
        bar.addWidget(self.filter_status)

        bar.addStretch()

        btn_export = QPushButton("📊 Xuất Excel doanh thu")
        btn_export.setStyleSheet(
            "QPushButton { background-color: #2b6cb0; color: white; border: none; "
            "border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background-color: #2c5282; }")
        btn_export.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_export.clicked.connect(self._on_export_revenue_excel)
        bar.addWidget(btn_export)

        btn_add = QPushButton("+ Tạo hóa đơn")
        btn_add.setStyleSheet(
            "QPushButton { background-color: #0b8480; color: white; border: none; "
            "border-radius: 8px; padding: 10px 20px; font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background-color: #096c69; }")
        btn_add.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_add.clicked.connect(self._on_add)
        bar.addWidget(btn_add)
        main_layout.addLayout(bar)

        # ── Table ──
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.HEADERS))
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(0, 80)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(9, 70)
        for col in range(1, 9):
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white; border: 1px solid #e2e8f0; border-radius: 8px;
                gridline-color: #edf2f7; font-size: 12px;
            }
            QTableWidget::item { padding: 4px; }
            QTableWidget::item:alternate { background-color: #f7fafc; }
            QHeaderView::section {
                background-color: #f7fafc; color: #718096; font-weight: bold;
                border: none; padding: 8px 4px; font-size: 11px;
            }
        """)
        main_layout.addWidget(self.table, 1)

        self.lbl_count = QLabel("Hiển thị 0 hóa đơn")
        self.lbl_count.setStyleSheet("color: #a0aec0; font-size: 12px;")
        main_layout.addWidget(self.lbl_count)

    # ── Data ──
    def _build_lookup(self):
        """Build guest/room lookup from contracts."""
        self._contracts = []
        self._guests_by_id = {}
        self._rooms_by_id = {}
        self._guests_rooms = {}  # contract_id -> (guest_name, room_number)

        if self.contract_service:
            try:
                self._contracts = self.contract_service.get_all()
            except Exception:
                pass
        if self.guest_service:
            try:
                for g in self.guest_service.get_all_guests():
                    self._guests_by_id[g.id] = g
            except Exception:
                pass
        if self.room_service:
            try:
                for r in self.room_service.get_all_rooms():
                    self._rooms_by_id[r.id] = r
            except Exception:
                pass

        for c in self._contracts:
            guest = self._guests_by_id.get(c.guest_id)
            room = self._rooms_by_id.get(c.room_id)
            self._guests_rooms[c.id] = (
                guest.full_name if guest else f"KT{c.guest_id:03d}",
                room.room_number if room else f"P{c.room_id}"
            )

    def refresh_data(self):
        if not self.invoice_service:
            return

        self._build_lookup()
        invoices = self.invoice_service.get_all()

        # Populate month filter from data (keep current selection)
        current_month_text = self.filter_month.currentText()
        self.filter_month.blockSignals(True)
        self.filter_month.clear()
        self.filter_month.addItem("Tất cả tháng")
        months_set = sorted(set((inv.month, inv.year) for inv in invoices),
                            key=lambda x: (x[1], x[0]), reverse=True)
        for m, y in months_set:
            self.filter_month.addItem(f"Tháng {m:02d}/{y}", (m, y))
        # Restore selection
        idx = self.filter_month.findText(current_month_text)
        if idx >= 0:
            self.filter_month.setCurrentIndex(idx)
        self.filter_month.blockSignals(False)

        # Search filter
        search_q = self.search_input.text().strip().lower()
        if search_q:
            invoices = [inv for inv in invoices if (
                search_q in inv.invoice_number.lower() or
                search_q in self._guests_rooms.get(inv.contract_id, ('', ''))[0].lower() or
                search_q in self._guests_rooms.get(inv.contract_id, ('', ''))[1].lower()
            )]

        # Month filter
        month_data = self.filter_month.currentData()
        if month_data:
            m, y = month_data
            invoices = [inv for inv in invoices if inv.month == m and inv.year == y]

        # Status filter
        status_map = {1: 'paid', 2: 'unpaid', 3: 'overdue'}
        status_f = status_map.get(self.filter_status.currentIndex())
        if status_f:
            if status_f == 'overdue':
                invoices = [inv for inv in invoices if inv.is_overdue()]
            else:
                invoices = [inv for inv in invoices if inv.status == status_f]

        # Stats
        all_inv = self.invoice_service.get_all()
        self.card_total.val_lbl.setText(str(len(all_inv)))
        self.card_paid.val_lbl.setText(str(sum(1 for i in all_inv if i.status == 'paid')))
        self.card_unpaid.val_lbl.setText(str(sum(1 for i in all_inv if i.status == 'unpaid')))
        from PyQt6.QtCore import QDate
        cur_month = QDate.currentDate().month()
        cur_year = QDate.currentDate().year()
        revenue = sum(i.total_amount for i in all_inv if i.status == 'paid' and i.month == cur_month and i.year == cur_year)
        if revenue >= 1_000_000:
            self.card_revenue.val_lbl.setText(f"{revenue/1_000_000:.1f}M")
        else:
            self.card_revenue.val_lbl.setText(f"{revenue:,.0f}đ")

        # Table
        self.table.setRowCount(len(invoices))
        for i, inv in enumerate(invoices):
            self.table.setRowHeight(i, 48)
            info = self._guests_rooms.get(inv.contract_id, ('—', '—'))

            self.table.setItem(i, 0, QTableWidgetItem(inv.invoice_number))
            # Guest name
            name_item = QTableWidgetItem(info[0])
            name_item.setToolTip(info[0])
            self.table.setItem(i, 1, name_item)
            self.table.setItem(i, 2, QTableWidgetItem(info[1]))
            self.table.setItem(i, 3, QTableWidgetItem(f"{inv.month:02d}/{inv.year}"))
            self.table.setItem(i, 4, QTableWidgetItem(f"{inv.room_rent:,.0f}đ"))
            self.table.setItem(i, 5, QTableWidgetItem(f"{inv.electricity_cost:,.0f}đ"))
            self.table.setItem(i, 6, QTableWidgetItem(f"{inv.water_cost:,.0f}đ"))

            # Total - bold
            total_item = QTableWidgetItem(f"{inv.total_amount:,.0f}đ")
            total_item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            total_item.setForeground(QColor("#2d3748"))
            self.table.setItem(i, 7, total_item)

            # Status
            st_text, st_color = self.STATUS_COLORS.get(
                'overdue' if inv.is_overdue() else inv.status, ('—', '#718096'))
            status_lbl = QLabel(st_text)
            status_lbl.setStyleSheet(f"color:{st_color};font-weight:bold;font-size:10px;background:transparent;")
            status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setCellWidget(i, 8, status_lbl)

            # Action
            actions = QWidget()
            actions.setStyleSheet("background: transparent;")
            al = QHBoxLayout(actions)
            al.setContentsMargins(2, 2, 2, 2)
            btn = QPushButton("Chi tiết")
            btn.setStyleSheet(
                "QPushButton{background:#ebf8ff;color:#3182ce;border:none;"
                "border-radius:4px;padding:4px 6px;font-weight:bold;font-size:10px;}"
                "QPushButton:hover{background:#bee3f8;}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(24)
            btn.clicked.connect(lambda _, inv=inv: self._on_detail(inv))
            al.addWidget(btn)
            self.table.setCellWidget(i, 9, actions)

        self.lbl_count.setText(f"Hiển thị {len(invoices)}/{len(all_inv)} hóa đơn")

    def _on_search(self, text):
        self.refresh_data()

    # ── CRUD ──
    def _on_add(self):
        self._build_lookup()
        active = [c for c in self._contracts if c.is_active()]
        dlg = InvoiceFormDialog(
            guests_rooms=self._guests_rooms, contracts=active, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.result_data
            inv = Invoice(
                invoice_number=f"HD{data['year']}{data['month']:02d}{len(self.invoice_service.get_all())+1:03d}",
                contract_id=data['contract_id'],
                month=data['month'], year=data['year'],
                room_rent=data['room_rent'],
                electricity_cost=data['electricity_cost'],
                water_cost=data['water_cost'],
                other_fees=data['other_fees'],
                total_amount=data['total_amount'],
                status='unpaid', due_date=data['due_date'],
            )
            self.invoice_service.invoice_repo.create(inv)
            show_success(self, "Thành công", "Tạo hóa đơn thành công")
            self.refresh_data()

    def _on_detail(self, inv: Invoice):
        info = self._guests_rooms.get(inv.contract_id, ('—', '—'))
        dlg = InvoiceDetailDialog(inv, extra={
            'guest_name': info[0], 'room_number': info[1]
        }, parent=self)
        dlg.delete_requested.connect(self._on_delete_confirmed)
        dlg.exec()

    def _on_pay(self, invoice_id: int):
        ok, msg = self.invoice_service.mark_paid(invoice_id)
        if ok:
            show_success(self, "Thành công", msg)
            self.refresh_data()
        else:
            show_warning(self, "Lỗi", msg)

    def _on_delete_confirmed(self, invoice_id: int):
        ok = self.invoice_service.invoice_repo.delete(invoice_id)
        if ok:
            show_success(self, "Thành công", "Xóa hóa đơn thành công")
            self.refresh_data()
        else:
            show_warning(self, "Lỗi", "Không thể xóa hóa đơn")

    def _on_export_revenue_excel(self):
        """Xuất báo cáo doanh thu ra Excel — chọn theo tháng."""
        from PyQt6.QtWidgets import QFileDialog, QFormLayout, QSpinBox
        from datetime import datetime
        from collections import defaultdict

        now = datetime.now()

        dlg = QDialog(self)
        dlg.setWindowTitle("Chọn khoảng thời gian xuất")
        dlg.setFixedSize(400, 240)
        dlg.setStyleSheet("QDialog { background: white; } QLabel { font-size: 13px; }")
        lay = QVBoxLayout(dlg)
        lay.setSpacing(12)
        lay.setContentsMargins(20, 20, 20, 20)
        title_lbl = QLabel("💰 Xuất báo cáo doanh thu")
        title_lbl.setStyleSheet("font-size: 15px; font-weight: bold; color: #2d3748;")
        lay.addWidget(title_lbl)
        cs = "QComboBox, QSpinBox { padding: 6px 10px; border: 1px solid #e2e8f0; border-radius: 6px; font-size: 13px; }"
        form = QFormLayout()
        form.setSpacing(10)
        rf = QHBoxLayout()
        month_from = QComboBox()
        month_from.addItems([f"Tháng {i}" for i in range(1, 13)])
        month_from.setCurrentIndex(0)
        month_from.setStyleSheet(cs)
        year_from = QSpinBox()
        year_from.setRange(2020, 2030)
        year_from.setValue(now.year)
        year_from.setStyleSheet(cs)
        rf.addWidget(month_from); rf.addWidget(year_from)
        form.addRow("Từ:", rf)
        rt = QHBoxLayout()
        month_to = QComboBox()
        month_to.addItems([f"Tháng {i}" for i in range(1, 13)])
        month_to.setCurrentIndex(now.month - 1)
        month_to.setStyleSheet(cs)
        year_to = QSpinBox()
        year_to.setRange(2020, 2030)
        year_to.setValue(now.year)
        year_to.setStyleSheet(cs)
        rt.addWidget(month_to); rt.addWidget(year_to)
        form.addRow("Đến:", rt)
        lay.addLayout(form)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        bc = QPushButton("Hủy")
        bc.setStyleSheet("QPushButton{background:#f7fafc;color:#4a5568;border:1px solid #cbd5e0;border-radius:8px;padding:8px 20px;font-weight:bold;}QPushButton:hover{background:#edf2f7;}")
        bc.clicked.connect(dlg.reject)
        btn_row.addWidget(bc)
        bo = QPushButton("Xuất Excel")
        bo.setStyleSheet("QPushButton{background-color:#2b6cb0;color:white;border:none;border-radius:8px;padding:8px 20px;font-weight:bold;}QPushButton:hover{background-color:#2c5282;}")
        bo.clicked.connect(dlg.accept)
        btn_row.addWidget(bo)
        lay.addLayout(btn_row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        m1, y1 = month_from.currentIndex() + 1, year_from.value()
        m2, y2 = month_to.currentIndex() + 1, year_to.value()

        # Không cho chọn tháng tương lai
        if (y2, m2) > (now.year, now.month):
            show_warning(self, "Lỗi", f"Chỉ được xuất tối đa đến tháng {now.month:02d}/{now.year}!")
            return
        if (y1, m1) > (y2, m2):
            show_warning(self, "Lỗi", "Tháng bắt đầu phải nhỏ hơn hoặc bằng tháng kết thúc!")
            return

        fp, _ = QFileDialog.getSaveFileName(self, "Lưu báo cáo doanh thu",
            f"DoanhThu_T{m1:02d}{y1}_T{m2:02d}{y2}.xlsx", "Excel Files (*.xlsx);;All Files (*)")
        if not fp:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Doanh thu"
            bd = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
            hf = Font(name='Arial', bold=True, color='FFFFFF', size=11)
            hfl = PatternFill(start_color='0B8480', end_color='0B8480', fill_type='solid')
            ha = Alignment(horizontal='center', vertical='center')
            df = Font(name='Arial', size=10)
            da = Alignment(vertical='center')
            ma = Alignment(horizontal='right', vertical='center')
            sub_f = Font(name='Arial', bold=True, size=10, color='2B6CB0')
            sub_fl = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
            mhf = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            mhfl = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')

            ws.merge_cells('A1:I1')
            tc = ws.cell(row=1, column=1, value=f"BÁO CÁO DOANH THU (T{m1:02d}/{y1} — T{m2:02d}/{y2})")
            tc.font = Font(name='Arial', bold=True, size=14, color='0B8480')
            tc.alignment = Alignment(horizontal='center')

            invoices = self.invoice_service.get_all() if self.invoice_service else []
            gmap, rmap, cmap = {}, {}, {}
            if self.guest_service:
                for g in self.guest_service.get_all_guests(): gmap[g.id] = g.full_name
            if self.room_service:
                for r in self.room_service.get_all_rooms(): rmap[r.id] = r.room_number
            if self.contract_service:
                for c in self.contract_service.get_all(): cmap[c.id] = c

            by_month = defaultdict(list)
            for inv in invoices:
                k = (inv.year, inv.month)
                if (y1, m1) <= k <= (y2, m2):
                    by_month[k].append(inv)

            sm = {'paid': 'Đã TT', 'unpaid': 'Chưa TT', 'overdue': 'Quá hạn'}
            hdrs = ['STT', 'Mã HĐ', 'Khách thuê', 'Phòng', 'Tiền phòng', 'Tiền điện', 'Tiền nước', 'Tổng tiền', 'Trạng thái']
            row = 3
            grand = 0

            for key in sorted(by_month.keys()):
                yr, mo = key
                mi = by_month[key]
                ws.merge_cells(f'A{row}:I{row}')
                mh = ws.cell(row=row, column=1, value=f"THÁNG {mo:02d}/{yr}")
                mh.font = mhf; mh.fill = mhfl; mh.alignment = ha; mh.border = bd
                row += 1
                for col, h in enumerate(hdrs, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = hf; c.fill = hfl; c.alignment = ha; c.border = bd
                row += 1

                mtot = 0
                for idx, inv in enumerate(mi, 1):
                    ct = cmap.get(inv.contract_id)
                    gn = gmap.get(inv.guest_id, '—') if hasattr(inv, 'guest_id') else '—'
                    rn = '—'
                    if ct:
                        gn = gmap.get(ct.guest_id, gn)
                        rn = rmap.get(ct.room_id, '—')
                    vals = [(idx, da), (inv.invoice_number, da), (gn, da), (rn, da),
                            (f"{inv.room_rent:,.0f}", ma), (f"{inv.electricity_cost:,.0f}", ma),
                            (f"{inv.water_cost:,.0f}", ma), (f"{inv.total_amount:,.0f}", ma),
                            (sm.get(inv.status, inv.status), da)]
                    for col, (v, a) in enumerate(vals, 1):
                        c = ws.cell(row=row, column=col, value=v)
                        c.font = df; c.alignment = a; c.border = bd
                    row += 1
                    if inv.status == 'paid':
                        mtot += inv.total_amount

                ws.merge_cells(f'A{row}:G{row}')
                st = ws.cell(row=row, column=1, value=f"Tổng tháng {mo:02d}/{yr}")
                st.font = sub_f; st.fill = sub_fl; st.alignment = Alignment(horizontal='right'); st.border = bd
                sv = ws.cell(row=row, column=8, value=f"{mtot:,.0f} VNĐ")
                sv.font = sub_f; sv.fill = sub_fl; sv.alignment = ma; sv.border = bd
                row += 2
                grand += mtot

            ws.merge_cells(f'A{row}:G{row}')
            gt = ws.cell(row=row, column=1, value='TỔNG DOANH THU ĐÃ THU')
            gt.font = Font(name='Arial', bold=True, size=13, color='0B8480')
            gt.alignment = Alignment(horizontal='right'); gt.border = bd
            gv = ws.cell(row=row, column=8, value=f"{grand:,.0f} VNĐ")
            gv.font = Font(name='Arial', bold=True, size=13, color='E53E3E')
            gv.alignment = ma; gv.border = bd

            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                ml = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value and not isinstance(cell, type(None)):
                        ml = max(ml, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(ml + 4, 30)

            wb.save(fp)
            show_success(self, "Xuất thành công", f"Đã lưu tại:\n{fp}")
            import os
            os.startfile(fp)
        except Exception as e:
            show_error(self, "Lỗi", f"Không thể xuất Excel:\n{e}")
